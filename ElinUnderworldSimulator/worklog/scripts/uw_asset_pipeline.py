"""Build importer-safe source sheets and canonical textures for ElinUnderworldSimulator."""

from __future__ import annotations

import argparse
from copy import copy
from functools import lru_cache
import os
import shutil
import tempfile
import zipfile
import xml.etree.ElementTree as ET

from openpyxl import Workbook, load_workbook
from PIL import Image, ImageDraw

from uw_asset_specs import (
    CHARA_COLUMNS,
    CHARA_ROWS,
    ELEMENT_COLUMNS,
    ELEMENT_ROWS,
    EXPECTED_TEXTURE_IDS,
    LANG_DIR,
    NUMERIC_FIELDS,
    OBJ_COLUMNS,
    OBJ_ROWS,
    PREVIEW_OUT,
    SOURCE_BLOCK_OUT,
    SOURCE_CARD_OUT,
    SOURCE_GAME_OUT,
    STAT_COLUMNS,
    STAT_ROWS,
    TEMPLATE_SOURCE_BLOCK,
    TEMPLATE_SOURCE_CARD,
    TEMPLATE_SOURCE_CHARA,
    TEMPLATE_SOURCE_GAME,
    TEXTURE_DIR,
    TEXTURE_SOURCE_MAP,
    THING_COLUMNS,
    THING_ROWS,
)

NS_MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
NS_REL = "http://schemas.openxmlformats.org/package/2006/relationships"
NS_CT = "http://schemas.openxmlformats.org/package/2006/content-types"
NS_DOC_REL = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"

ET.register_namespace("", NS_MAIN)
ET.register_namespace("r", NS_DOC_REL)

EXPECTED_STATION_SKILL_ALIASES = {
    "uw_mixing_table": "handicraft",
    "uw_advanced_lab": "handicraft",
}


def ensure_dirs() -> None:
    LANG_DIR.mkdir(parents=True, exist_ok=True)
    TEXTURE_DIR.mkdir(parents=True, exist_ok=True)


def copy_template_rows(src_sheet, dst_sheet, row_count: int, col_count: int) -> None:
    for row_idx in range(1, row_count + 1):
        dst_sheet.row_dimensions[row_idx].height = src_sheet.row_dimensions[row_idx].height
        for col_idx in range(1, col_count + 1):
            src_cell = src_sheet.cell(row=row_idx, column=col_idx)
            dst_cell = dst_sheet.cell(row=row_idx, column=col_idx)
            dst_cell.value = src_cell.value
            if src_cell.has_style:
                dst_cell._style = copy(src_cell._style)
            if src_cell.number_format:
                dst_cell.number_format = src_cell.number_format
            if src_cell.font:
                dst_cell.font = copy(src_cell.font)
            if src_cell.fill:
                dst_cell.fill = copy(src_cell.fill)
            if src_cell.border:
                dst_cell.border = copy(src_cell.border)
            if src_cell.alignment:
                dst_cell.alignment = copy(src_cell.alignment)
            if src_cell.protection:
                dst_cell.protection = copy(src_cell.protection)

    for col_letter, dim in src_sheet.column_dimensions.items():
        dst_sheet.column_dimensions[col_letter].width = dim.width
        dst_sheet.column_dimensions[col_letter].hidden = dim.hidden

    for merged in src_sheet.merged_cells.ranges:
        dst_sheet.merge_cells(str(merged))

    dst_sheet.freeze_panes = src_sheet.freeze_panes
    dst_sheet.sheet_view.zoomScale = src_sheet.sheet_view.zoomScale


def write_rows(sheet, columns: list[str], rows: list[dict[str, object]], start_row: int = 4) -> None:
    for row_offset, data in enumerate(rows, start=start_row):
        for col_idx, column_name in enumerate(columns, start=1):
            if column_name not in data:
                continue
            value = data[column_name]
            if value is None or value == "":
                continue
            sheet.cell(row=row_offset, column=col_idx).value = value


def read_shared_strings(zin: zipfile.ZipFile) -> list[str]:
    if "xl/sharedStrings.xml" not in zin.namelist():
        return []
    root = ET.fromstring(zin.read("xl/sharedStrings.xml"))
    values: list[str] = []
    for si in root.findall(f"{{{NS_MAIN}}}si"):
        values.append("".join(text_node.text or "" for text_node in si.iter(f"{{{NS_MAIN}}}t")))
    return values


def normalize_shared_strings(path) -> None:
    shared: list[str] = []
    shared_index: dict[str, int] = {}

    def intern(text: str) -> int:
        if text not in shared_index:
            shared_index[text] = len(shared)
            shared.append(text)
        return shared_index[text]

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        tmp_path = tmp.name

    try:
        with zipfile.ZipFile(path, "r") as zin:
            old_shared = read_shared_strings(zin)
            with zipfile.ZipFile(tmp_path, "w", zipfile.ZIP_DEFLATED) as zout:
                for info in zin.infolist():
                    if info.filename == "xl/sharedStrings.xml":
                        continue

                    data = zin.read(info.filename)
                    if info.filename.startswith("xl/worksheets/sheet") and info.filename.endswith(".xml"):
                        root = ET.fromstring(data)
                        for cell in root.iter(f"{{{NS_MAIN}}}c"):
                            cell_type = cell.get("t")
                            if cell_type not in ("inlineStr", "s", "str"):
                                continue

                            text = ""
                            if cell_type == "inlineStr":
                                inline = cell.find(f"{{{NS_MAIN}}}is")
                                if inline is not None:
                                    text = "".join(node.text or "" for node in inline.iter(f"{{{NS_MAIN}}}t"))
                            else:
                                value_node = cell.find(f"{{{NS_MAIN}}}v")
                                if value_node is not None:
                                    if cell_type == "s":
                                        idx = int(value_node.text or "0")
                                        text = old_shared[idx] if 0 <= idx < len(old_shared) else ""
                                    else:
                                        text = value_node.text or ""

                            for child in list(cell):
                                cell.remove(child)
                            cell.set("t", "s")
                            value_node = ET.SubElement(cell, f"{{{NS_MAIN}}}v")
                            value_node.text = str(intern(text))
                        data = ET.tostring(root, encoding="utf-8", xml_declaration=True)

                    elif info.filename == "[Content_Types].xml":
                        root = ET.fromstring(data)
                        part_name = "/xl/sharedStrings.xml"
                        exists = any(
                            child.tag == f"{{{NS_CT}}}Override" and child.get("PartName") == part_name
                            for child in root
                        )
                        if not exists:
                            ET.SubElement(
                                root,
                                f"{{{NS_CT}}}Override",
                                {
                                    "PartName": part_name,
                                    "ContentType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sharedStrings+xml",
                                },
                            )
                        data = ET.tostring(root, encoding="utf-8", xml_declaration=True)

                    elif info.filename == "xl/_rels/workbook.xml.rels":
                        root = ET.fromstring(data)
                        exists = any(
                            child.tag == f"{{{NS_REL}}}Relationship"
                            and child.get("Type") == f"{NS_DOC_REL}/sharedStrings"
                            for child in root
                        )
                        if not exists:
                            used_ids = {
                                int((child.get("Id") or "rId0").replace("rId", ""))
                                for child in root
                                if (child.get("Id") or "").startswith("rId")
                            }
                            next_id = 1
                            while next_id in used_ids:
                                next_id += 1
                            ET.SubElement(
                                root,
                                f"{{{NS_REL}}}Relationship",
                                {
                                    "Id": f"rId{next_id}",
                                    "Type": f"{NS_DOC_REL}/sharedStrings",
                                    "Target": "sharedStrings.xml",
                                },
                            )
                        data = ET.tostring(root, encoding="utf-8", xml_declaration=True)

                    zout.writestr(info, data)

                sst = ET.Element(f"{{{NS_MAIN}}}sst", {"count": str(len(shared)), "uniqueCount": str(len(shared))})
                for text in shared:
                    si = ET.SubElement(sst, f"{{{NS_MAIN}}}si")
                    t = ET.SubElement(si, f"{{{NS_MAIN}}}t")
                    if text != text.strip():
                        t.set("{http://www.w3.org/XML/1998/namespace}space", "preserve")
                    t.text = text
                zout.writestr("xl/sharedStrings.xml", ET.tostring(sst, encoding="utf-8", xml_declaration=True))

        shutil.move(tmp_path, path)
    finally:
        if tmp_path and os.path.exists(tmp_path):
            os.unlink(tmp_path)


def build_source_card() -> None:
    ensure_dirs()
    thing_template_wb = load_workbook(TEMPLATE_SOURCE_CARD)
    chara_template_wb = load_workbook(TEMPLATE_SOURCE_CHARA)

    workbook = Workbook()
    workbook.remove(workbook.active)

    thing_sheet = workbook.create_sheet("Thing")
    chara_sheet = workbook.create_sheet("Chara")

    copy_template_rows(thing_template_wb["Thing"], thing_sheet, row_count=3, col_count=len(THING_COLUMNS))
    copy_template_rows(chara_template_wb["Chara"], chara_sheet, row_count=3, col_count=len(CHARA_COLUMNS))
    write_rows(thing_sheet, THING_COLUMNS, THING_ROWS)
    write_rows(chara_sheet, CHARA_COLUMNS, CHARA_ROWS)

    workbook.save(SOURCE_CARD_OUT)
    normalize_shared_strings(SOURCE_CARD_OUT)


def clone_template_workbook(template_path, output_path, sheet_rows: dict[str, tuple[list[str], list[dict[str, object]], int]]) -> None:
    ensure_dirs()
    template_wb = load_workbook(template_path)

    workbook = Workbook()
    workbook.remove(workbook.active)

    for sheet_name in template_wb.sheetnames:
        src_sheet = template_wb[sheet_name]
        dst_sheet = workbook.create_sheet(sheet_name)
        col_count = src_sheet.max_column
        copy_template_rows(src_sheet, dst_sheet, row_count=3, col_count=col_count)
        if sheet_name in sheet_rows:
            columns, rows, start_row = sheet_rows[sheet_name]
            write_rows(dst_sheet, columns, rows, start_row=start_row)

    workbook.save(output_path)
    normalize_shared_strings(output_path)


def build_source_block() -> None:
    clone_template_workbook(
        TEMPLATE_SOURCE_BLOCK,
        SOURCE_BLOCK_OUT,
        {"Obj": (OBJ_COLUMNS, OBJ_ROWS, 4)},
    )


def build_source_game() -> None:
    clone_template_workbook(
        TEMPLATE_SOURCE_GAME,
        SOURCE_GAME_OUT,
        {
            "Element": (ELEMENT_COLUMNS, ELEMENT_ROWS, 4),
            "Stat": (STAT_COLUMNS, STAT_ROWS, 11),
        },
    )


def materialize_textures() -> None:
    ensure_dirs()
    missing_textures: list[str] = []
    for asset_id in TEXTURE_SOURCE_MAP:
        dest_path = TEXTURE_DIR / f"{asset_id}.png"
        if not dest_path.exists():
            missing_textures.append(asset_id)

    if missing_textures:
        raise FileNotFoundError(
            "Missing required textures: " + ", ".join(sorted(missing_textures))
        )


def build_preview() -> None:
    ensure_dirs()
    preview_ids = [
        "uw_mixing_table",
        "uw_processing_vat",
        "uw_advanced_lab",
        "uw_herb_whisper",
        "uw_powder_dream",
        "uw_elixir_shadow",
        "uw_incense_ash",
        "uw_tonic_whisper_refined",
        "uw_fixer",
    ]

    tiles = []
    for asset_id in preview_ids:
        image = Image.open(TEXTURE_DIR / f"{asset_id}.png").convert("RGBA")
        tile = Image.new("RGBA", (180, 120), (18, 22, 28, 255))
        image.thumbnail((96, 96), Image.Resampling.NEAREST)
        tile.alpha_composite(image, ((tile.width - image.width) // 2, 8))
        draw = ImageDraw.Draw(tile)
        draw.rounded_rectangle((0, 0, tile.width - 1, tile.height - 1), radius=12, outline=(90, 190, 150, 255), width=2)
        draw.text((12, 97), asset_id.replace("uw_", "").replace("_", " "), fill=(234, 241, 228, 255))
        tiles.append(tile)

    preview = Image.new("RGB", (960, 640), (12, 16, 22))
    draw = ImageDraw.Draw(preview)
    draw.text((36, 28), "Elin Underworld Simulator", fill=(237, 242, 232))
    draw.text((36, 62), "Crafting, farming, smoking core slice", fill=(130, 208, 176))
    for idx, tile in enumerate(tiles):
        row = idx // 3
        col = idx % 3
        preview.paste(tile.convert("RGB"), (36 + col * 204, 118 + row * 150))
    preview.save(PREVIEW_OUT, quality=92)


def workbook_sheet_xml_entries(path) -> list[str]:
    with zipfile.ZipFile(path, "r") as archive:
        return sorted(name for name in archive.namelist() if name.startswith("xl/worksheets/") and name.endswith(".xml"))


def assert_no_inline_strings(path) -> None:
    with zipfile.ZipFile(path, "r") as archive:
        if "xl/sharedStrings.xml" not in archive.namelist():
            raise ValueError(f"{path.name} is missing xl/sharedStrings.xml")
        for entry_name in sorted(
            name for name in archive.namelist() if name.startswith("xl/worksheets/") and name.endswith(".xml")
        ):
            text = archive.read(entry_name).decode("utf-8")
            if 't="inlineStr"' in text or "<is>" in text:
                raise ValueError(f"{path.name} still contains inline strings in {entry_name}")


@lru_cache(maxsize=1)
def _template_header_snapshot() -> tuple[tuple[tuple[object, ...], ...], tuple[tuple[object, ...], ...]]:
    template_card = load_workbook(TEMPLATE_SOURCE_CARD, read_only=True, data_only=False)
    template_chara = load_workbook(TEMPLATE_SOURCE_CHARA, read_only=True, data_only=False)
    try:
        return (
            tuple(
                tuple(row)
                for row in template_card["Thing"].iter_rows(
                    min_row=1,
                    max_row=3,
                    min_col=1,
                    max_col=len(THING_COLUMNS),
                    values_only=True,
                )
            ),
            tuple(
                tuple(row)
                for row in template_chara["Chara"].iter_rows(
                    min_row=1,
                    max_row=3,
                    min_col=1,
                    max_col=len(CHARA_COLUMNS),
                    values_only=True,
                )
            ),
        )
    finally:
        template_card.close()
        template_chara.close()


def assert_template_rows_preserved(generated_card=None) -> None:
    close_generated = generated_card is None
    if generated_card is None:
        generated_card = load_workbook(SOURCE_CARD_OUT, read_only=True, data_only=False)

    try:
        expected_thing_rows, expected_chara_rows = _template_header_snapshot()
        actual_thing_rows = tuple(
            tuple(row)
            for row in generated_card["Thing"].iter_rows(
                min_row=1,
                max_row=3,
                min_col=1,
                max_col=len(THING_COLUMNS),
                values_only=True,
            )
        )
        actual_chara_rows = tuple(
            tuple(row)
            for row in generated_card["Chara"].iter_rows(
                min_row=1,
                max_row=3,
                min_col=1,
                max_col=len(CHARA_COLUMNS),
                values_only=True,
            )
        )

        if actual_thing_rows != expected_thing_rows:
            raise ValueError("Thing sheet template rows were modified.")

        if actual_chara_rows != expected_chara_rows:
            raise ValueError("Chara sheet template rows were modified.")
    finally:
        if close_generated:
            generated_card.close()


def assert_numeric_integrity(workbook, path, sheet_name: str, columns: list[str], start_row: int, row_count: int) -> None:
    sheet = workbook[sheet_name]
    numeric_fields = NUMERIC_FIELDS.get(sheet_name, set())
    numeric_columns = [(columns.index(name) + 1, name) for name in columns if name in numeric_fields]
    end_row = start_row + row_count - 1
    for row_idx, row in enumerate(
        sheet.iter_rows(
            min_row=start_row,
            max_row=end_row,
            min_col=1,
            max_col=len(columns),
            values_only=True,
        ),
        start=start_row,
    ):
        for col_idx, column_name in numeric_columns:
            value = row[col_idx - 1]
            if value in (None, ""):
                continue
            if not isinstance(value, (int, float)):
                raise ValueError(f"{path.name}:{sheet_name} row {row_idx} column {column_name} must be numeric, got {value!r}")


def parse_components(components: str | None) -> list[str]:
    if not components:
        return []
    result: list[str] = []
    for part in components.split(","):
        item = part.strip()
        if not item or item == "-":
            continue
        item_id = item.split("/")[0].split("|")[0].split("@")[0].strip()
        if item_id.startswith("#"):
            continue
        result.append(item_id)
    return result


@lru_cache(maxsize=1)
def build_valid_base_ids() -> set[str]:
    workbook = load_workbook(TEMPLATE_SOURCE_CARD, read_only=True, data_only=False)
    try:
        sheet = workbook["Thing"]
        ids = set()
        for (value,) in sheet.iter_rows(min_row=4, min_col=1, max_col=1, values_only=True):
            if isinstance(value, str) and value:
                ids.add(value)
        return ids
    finally:
        workbook.close()


def assert_recipe_and_factory_integrity() -> None:
    known_ids = build_valid_base_ids() | {row["id"] for row in THING_ROWS}
    known_factories = known_ids | {"self", "workbench"}
    known_conditions = {row["alias"] for row in STAT_ROWS}

    for row in THING_ROWS:
        for component_id in parse_components(row.get("components")):
            if component_id not in known_ids:
                raise ValueError(f"Unknown recipe component '{component_id}' in {row['id']}")

        factory = row.get("factory")
        if factory and factory not in known_factories:
            raise ValueError(f"Unknown factory '{factory}' in {row['id']}")

        trait = row.get("trait")
        if isinstance(trait, str):
            for token in trait.split(","):
                token = token.strip()
                if token.startswith("ConUW") and token not in known_conditions:
                    raise ValueError(f"Unknown condition '{token}' referenced by {row['id']}")


def assert_station_skill_aliases() -> None:
    for row in THING_ROWS:
        expected_alias = EXPECTED_STATION_SKILL_ALIASES.get(row["id"])
        if expected_alias is None:
            continue

        trait = row.get("trait")
        if not isinstance(trait, str):
            raise ValueError(f"{row['id']} is missing required trait metadata.")

        tokens = [token.strip() for token in trait.split(",") if token.strip()]
        if len(tokens) < 2:
            raise ValueError(f"{row['id']} must declare a skill alias in its trait string.")

        actual_alias = tokens[1]
        if actual_alias != expected_alias:
            raise ValueError(
                f"{row['id']} must use vanilla skill alias '{expected_alias}', got '{actual_alias}'."
            )

def assert_texture_parity() -> None:
    actual = {path.stem for path in TEXTURE_DIR.glob("*.png")}
    missing = sorted(EXPECTED_TEXTURE_IDS - actual)
    if missing:
        raise ValueError(f"Missing canonical textures: {', '.join(missing)}")


def verify() -> None:
    for path in (SOURCE_CARD_OUT, SOURCE_BLOCK_OUT, SOURCE_GAME_OUT):
        if not path.exists():
            raise FileNotFoundError(f"Missing workbook: {path}")
        assert_no_inline_strings(path)

    generated_card = load_workbook(SOURCE_CARD_OUT, read_only=True, data_only=False)
    generated_block = load_workbook(SOURCE_BLOCK_OUT, read_only=True, data_only=False)
    generated_game = load_workbook(SOURCE_GAME_OUT, read_only=True, data_only=False)
    try:
        assert_template_rows_preserved(generated_card)
        assert_numeric_integrity(generated_card, SOURCE_CARD_OUT, "Thing", THING_COLUMNS, 4, len(THING_ROWS))
        assert_numeric_integrity(generated_card, SOURCE_CARD_OUT, "Chara", CHARA_COLUMNS, 4, len(CHARA_ROWS))
        assert_numeric_integrity(generated_block, SOURCE_BLOCK_OUT, "Obj", OBJ_COLUMNS, 4, len(OBJ_ROWS))
        assert_numeric_integrity(generated_game, SOURCE_GAME_OUT, "Element", ELEMENT_COLUMNS, 4, len(ELEMENT_ROWS))
        assert_numeric_integrity(generated_game, SOURCE_GAME_OUT, "Stat", STAT_COLUMNS, 11, len(STAT_ROWS))
        assert_recipe_and_factory_integrity()
        assert_station_skill_aliases()
        assert_texture_parity()
    finally:
        generated_card.close()
        generated_block.close()
        generated_game.close()


def build_all() -> None:
    materialize_textures()
    build_source_card()
    build_source_block()
    build_source_game()
    build_preview()


def run_integrate(only_ids: list[str] | None = None) -> None:
    from integrate_assets import integrate_assets

    integrate_assets(only_ids=only_ids)
    build_all()
    verify()


def run_generate(only_ids: list[str] | None = None, model: str = "nano-banana-pro-preview") -> None:
    from generate_assets import run_generation

    run_generation(only_ids=only_ids, model=model)


def main() -> None:
    parser = argparse.ArgumentParser(description="Underworld asset/source-sheet pipeline")
    parser.add_argument("command", choices=("build", "verify", "all", "generate", "integrate"), nargs="?", default="all")
    parser.add_argument("--only", nargs="+", help="Limit generation/integration to specific asset ids")
    parser.add_argument("--model", default="nano-banana-pro-preview", help="Gemini image model for generation")
    args = parser.parse_args()

    if args.command == "generate":
        run_generate(only_ids=args.only, model=args.model)
    elif args.command == "integrate":
        run_integrate(only_ids=args.only)
    elif args.command == "verify":
        verify()
    elif args.command == "build":
        build_all()
    else:
        build_all()
        verify()

    print("Underworld asset pipeline completed successfully.")


if __name__ == "__main__":
    main()
