"""Ensure Skyreader Guild source rows are present and importer-safe.

Elin's NPOI-based importer expects workbook strings through sharedStrings.xml.
openpyxl writes inlineStr cells by default, which Elin imports as empty values.
This script validates the mod's key Thing rows, fixes known data drift, and
normalizes workbook strings after saving.
"""

import io
import shutil
import sys
import tempfile
import zipfile
import xml.etree.ElementTree as ET
import json
import re
import os
from pathlib import Path

import openpyxl

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

ROOT = Path(__file__).resolve().parent.parent.parent
SOURCE_CARD = ROOT / "LangMod" / "EN" / "SourceCard.xlsx"

NS_MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
NS_REL = "http://schemas.openxmlformats.org/package/2006/relationships"
NS_CT = "http://schemas.openxmlformats.org/package/2006/content-types"
NS_DOC_REL = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"

ET.register_namespace("", NS_MAIN)
ET.register_namespace("r", NS_DOC_REL)


def read_shared_strings(zin):
    if "xl/sharedStrings.xml" not in zin.namelist():
        return []
    root = ET.fromstring(zin.read("xl/sharedStrings.xml"))
    values = []
    for si in root.findall(f"{{{NS_MAIN}}}si"):
        parts = []
        for t in si.iter(f"{{{NS_MAIN}}}t"):
            parts.append(t.text or "")
        values.append("".join(parts))
    return values


def normalize_shared_strings(path):
    """Convert all string cells to shared strings for Elin's source importer."""
    shared = []
    shared_index = {}

    def intern(text):
        if text not in shared_index:
            shared_index[text] = len(shared)
            shared.append(text)
        return shared_index[text]

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        tmp_path = Path(tmp.name)

    try:
        with zipfile.ZipFile(path, "r") as zin:
            old_shared = read_shared_strings(zin)
            with zipfile.ZipFile(tmp_path, "w", zipfile.ZIP_DEFLATED) as zout:
                for info in zin.infolist():
                    if info.filename == "xl/sharedStrings.xml":
                        continue

                    data = zin.read(info.filename)
                    if info.filename.startswith("xl/worksheets/sheet") and info.filename.endswith(".xml"):
                        root = ET.fromstring(data)
                        for cell in root.iter(f"{{{NS_MAIN}}}c"):
                            cell_type = cell.get("t")
                            if cell_type not in ("inlineStr", "s", "str"):
                                continue

                            text = ""
                            if cell_type == "inlineStr":
                                inline = cell.find(f"{{{NS_MAIN}}}is")
                                if inline is not None:
                                    text = "".join(t.text or "" for t in inline.iter(f"{{{NS_MAIN}}}t"))
                            else:
                                v = cell.find(f"{{{NS_MAIN}}}v")
                                if v is not None:
                                    if cell_type == "s":
                                        idx = int(v.text or "0")
                                        text = old_shared[idx] if 0 <= idx < len(old_shared) else ""
                                    else:
                                        text = v.text or ""

                            for child in list(cell):
                                cell.remove(child)
                            cell.set("t", "s")
                            v = ET.SubElement(cell, f"{{{NS_MAIN}}}v")
                            v.text = str(intern(text))
                        data = ET.tostring(root, encoding="utf-8", xml_declaration=True)

                    elif info.filename == "[Content_Types].xml":
                        root = ET.fromstring(data)
                        part_name = "/xl/sharedStrings.xml"
                        exists = any(
                            child.tag == f"{{{NS_CT}}}Override" and child.get("PartName") == part_name
                            for child in root
                        )
                        if not exists:
                            ET.SubElement(
                                root,
                                f"{{{NS_CT}}}Override",
                                {
                                    "PartName": part_name,
                                    "ContentType": (
                                        "application/vnd.openxmlformats-officedocument."
                                        "spreadsheetml.sharedStrings+xml"
                                    ),
                                },
                            )
                        data = ET.tostring(root, encoding="utf-8", xml_declaration=True)

                    elif info.filename == "xl/_rels/workbook.xml.rels":
                        root = ET.fromstring(data)
                        exists = any(
                            child.tag == f"{{{NS_REL}}}Relationship"
                            and child.get("Type") == f"{NS_DOC_REL}/sharedStrings"
                            for child in root
                        )
                        if not exists:
                            used = {
                                int((child.get("Id") or "rId0").replace("rId", ""))
                                for child in root
                                if (child.get("Id") or "").startswith("rId")
                            }
                            next_id = 1
                            while next_id in used:
                                next_id += 1
                            ET.SubElement(
                                root,
                                f"{{{NS_REL}}}Relationship",
                                {
                                    "Id": f"rId{next_id}",
                                    "Type": f"{NS_DOC_REL}/sharedStrings",
                                    "Target": "sharedStrings.xml",
                                },
                            )
                        data = ET.tostring(root, encoding="utf-8", xml_declaration=True)

                    zout.writestr(info, data)

                sst = ET.Element(
                    f"{{{NS_MAIN}}}sst",
                    {"count": str(len(shared)), "uniqueCount": str(len(shared))},
                )
                for text in shared:
                    si = ET.SubElement(sst, f"{{{NS_MAIN}}}si")
                    t = ET.SubElement(si, f"{{{NS_MAIN}}}t")
                    if text != text.strip():
                        t.set("{http://www.w3.org/XML/1998/namespace}space", "preserve")
                    t.text = text
                zout.writestr(
                    "xl/sharedStrings.xml",
                    ET.tostring(sst, encoding="utf-8", xml_declaration=True),
                )

        shutil.move(tmp_path, path)
    finally:
        if tmp_path.exists():
            tmp_path.unlink()


THING_COL = {
    "id": 1,
    "name_JP": 2,
    "unit_JP": 4,
    "name": 6,
    "category": 9,
    "sort": 10,
    "sort_value": 11,
    "_tileType": 12,
    "_idRenderData": 13,
    "tiles": 14,
    "altTiles": 15,
    "anime": 16,
    "recipeKey": 21,
    "factory": 22,
    "components": 23,
    "defMat": 25,
    "value": 27,
    "LV": 28,
    "chance": 29,
    "quality": 30,
    "HP": 31,
    "weight": 32,
    "trait": 34,
    "elements": 35,
    "lightData": 41,
    "idExtra": 42,
    "idToggleExtra": 43,
    "idSound": 45,
    "tag": 46,
    "roomName_JP": 49,
    "roomName": 50,
    "detail_JP": 51,
    "detail": 52,
}


CHARA_COL = {
    "id": 1,
    "_id": 2,
    "name_JP": 3,
    "name": 4,
    "aka_JP": 5,
    "aka": 6,
    "idActor": 7,
    "sort": 8,
    "size": 9,
    "_idRenderData": 10,
    "tiles": 11,
    "tiles_snow": 12,
    "colorMod": 13,
    "components": 14,
    "defMat": 15,
    "LV": 16,
    "chance": 17,
    "quality": 18,
    "hostility": 19,
    "biome": 20,
    "tag": 21,
    "trait": 22,
    "race": 23,
    "job": 24,
    "tactics": 25,
    "aiIdle": 26,
    "aiParam": 27,
    "actCombat": 28,
    "mainElement": 29,
    "elements": 30,
    "equip": 31,
    "loot": 32,
    "category": 33,
    "filter": 34,
    "gachaFilter": 35,
    "tone": 36,
    "actIdle": 37,
    "lightData": 38,
    "idExtra": 39,
    "bio": 40,
    "faith": 41,
    "works": 42,
    "hobbies": 43,
    "idText": 44,
    "moveAnime": 45,
    "factory": 46,
    "components_2": 47,
    "recruitItems": 48,
    "detail_JP": 49,
    "detail": 50,
}


FORBIDDEN_CHARA_IDS = {
    "chickchicken",
    "chickchick",
    "chickchik",
    "test_chicken",
}


EXPECTED_THINGS = {
    "srg_starchart": {
        "name_JP": "淡い星図",
        "unit_JP": "冊",
        "name": "faint starchart",
        "category": "map",
        "sort": "book",
        "_idRenderData": "@obj_S flat",
        "tiles": 1611,
        "components": "texture",
        "defMat": "paper",
        "value": 100,
        "LV": 1,
        "weight": 150,
        "trait": "ScrollMap",
        "detail": "The text on the page seems to waver faintly...",
    },
    "srg_codex": {
        "name_JP": "占星術写本",
        "name": "astrological codex",
        "category": "crafter",
        "_tileType": "ObjBig",
        "_idRenderData": "@obj",
        "tiles": 1552,
        "recipeKey": "*",
        "factory": "workbench",
        "components": "srg_meteorite_source/4,ore_gem/12,log/4,ingot/9",
        "defMat": "paper",
        "value": 5000,
        "LV": 10,
        "chance": 0,
        "HP": 80,
        "weight": 12000,
        "trait": "AstrologicalCodex,reading",
        "detail_JP": "A codex and instruments of strange design.",
        "detail": "A codex and instruments of strange design.",
    },
    "srg_astral_extractor": {
        "name_JP": "星霊抽出器",
        "name": "astral extractor",
        "category": "_item",
        "_idRenderData": "@obj_S flat",
        "tiles": 1208,
        "recipeKey": "*",
        "factory": "srg_codex",
        "components": "srg_meteorite_source/1,potion_empty/1",
        "defMat": "glass",
        "value": 300,
        "LV": 1,
        "chance": 0,
        "weight": 200,
        "trait": "AstralExtractor",
        "detail_JP": "A device that pulls stray starlight from meteor-touched beings and objects.",
        "detail": "A device that pulls stray starlight from meteor-touched beings and objects.",
    },
    "srg_meteor_core": {
        "name_JP": "隕石の核",
        "name": "meteor core",
        "category": "junk",
        "sort": "junk",
        "_idRenderData": "@obj_S",
        "tiles": 503,
        "defMat": "granite",
        "value": 500,
        "LV": 1,
        "weight": 5000,
        "trait": "MeteorCore",
        "detail": "A pulsing core of extraterrestrial origin. It thrums with energy.",
    },
    "srg_meteorite_source": {
        "name_JP": "隕石素材",
        "name": "meteorite source",
        "category": "ore",
        "sort": "resource_ore",
        "_idRenderData": "@obj_S flat",
        "tiles": 530,
        "defMat": "gold",
        "value": 200,
        "LV": 1,
        "weight": 500,
        "trait": "ResourceMain",
        "detail": "A fragment of fallen star, rich with rare minerals.",
    },
    "srg_debris": {
        "name_JP": "衝突破片",
        "name": "impact debris",
        "category": "junk",
        "sort": "junk",
        "_idRenderData": "@obj_S",
        "tiles": 503,
        "defMat": "granite",
        "value": 10,
        "LV": 1,
        "weight": 2000,
        "trait": "SrgDebris",
        "detail": "Scorched fragments from a meteor impact.",
    },
    "srg_weave_stars": {
        "name_JP": "星の織物",
        "name": "weave of the stars",
        "category": "_item",
        "_idRenderData": "@obj_S flat",
        "tiles": 504,
        "recipeKey": "*",
        "factory": "srg_codex",
        "components": "srg_meteorite_source/2",
        "defMat": "gold",
        "value": 1000,
        "LV": 1,
        "chance": 0,
        "weight": 100,
        "trait": "StarImbuement,weave",
        "detail_JP": "Starlight folded into a form that can settle into armor.",
        "detail": "Starlight folded into a form that can settle into armor.",
    },
    "srg_starforge": {
        "name_JP": "星鍛の火花",
        "name": "starforge spark",
        "category": "_item",
        "_idRenderData": "@obj_S flat",
        "tiles": 530,
        "recipeKey": "*",
        "factory": "srg_codex",
        "components": "srg_meteorite_source/2",
        "defMat": "gold",
        "value": 1200,
        "LV": 1,
        "chance": 0,
        "weight": 100,
        "trait": "StarImbuement,forge",
        "detail_JP": "A shard of intent for awakening weapons beneath starlight.",
        "detail": "A shard of intent for awakening weapons beneath starlight.",
    },
    "srg_astral_portal": {
        "name_JP": "星霊の門",
        "name": "astral portal",
        "category": "_item",
        "_idRenderData": "@obj tall",
        "tiles": 554,
        "anime": "4,500",
        "defMat": "glass",
        "value": 0,
        "LV": 1,
        "weight": 99999,
        "trait": "AstralPortal",
        "tag": "noShop,noWish",
        "detail_JP": "A shimmering gateway to an astral rift.",
        "detail": "A shimmering gateway to an astral rift.",
    },
    "srg_guild_entrance": {
        "name_JP": "",
        "name": "Skyreader Observatory gateway",
        "category": "mech",
        "sort": "furniture_mech",
        "_idRenderData": "obj tall",
        "tiles": 751,
        "altTiles": 752,
        "anime": "4,500",
        "defMat": "ether",
        "value": 0,
        "LV": 1,
        "chance": 0,
        "quality": 4,
        "weight": 10000,
        "trait": "SkyreaderGuildEntrance,srg_guild_hq",
        "tag": "noWish,noShop",
        "roomName": "Portal Room",
        "detail_JP": "",
        "detail": "A shimmering astral gate leading to the Skyreader Observatory.",
    },
    "srg_guild_exit": {
        "name_JP": "",
        "name": "Derphy gateway",
        "category": "mech",
        "sort": "furniture_mech",
        "_idRenderData": "obj tall",
        "tiles": 751,
        "altTiles": 752,
        "anime": "4,500",
        "defMat": "ether",
        "value": 0,
        "LV": 1,
        "chance": 0,
        "quality": 4,
        "weight": 10000,
        "trait": "SkyreaderGuildExit",
        "tag": "noWish,noShop",
        "roomName": "Portal Room",
        "detail_JP": "",
        "detail": "A shimmering astral gate leading back to Derphy.",
    },
    "srg_ladder_plaque": {
        "name_JP": "",
        "name": "Starlight Ladder plaque",
        "category": "deco",
        "_tileType": "",
        "_idRenderData": "@obj tall",
        "tiles": 1552,
        "defMat": "gold",
        "value": 1600,
        "LV": 1,
        "chance": 0,
        "quality": 3,
        "weight": 2800,
        "trait": "LadderPlaque",
        "tag": "noShop,noWish",
        "roomName": "Starlight Ladder",
        "detail_JP": "",
        "detail": "A guild plaque covered with carefully updated rankings.",
    },
    "srg_scroll_twilight": {
        "category": "scroll",
        "sort": "book_scroll",
        "components": "srg_meteorite_source/5,zettel/1",
        "trait": "BossScroll",
        "tag": "noShop",
    },
    "srg_scroll_radiance": {
        "category": "scroll",
        "sort": "book_scroll",
        "components": "srg_meteorite_source/5,zettel/1",
        "trait": "BossScroll",
        "tag": "noShop",
    },
    "srg_scroll_abyss": {
        "category": "scroll",
        "sort": "book_scroll",
        "components": "srg_meteorite_source/5,zettel/1",
        "trait": "BossScroll",
        "tag": "noShop",
    },
    "srg_scroll_nova": {
        "category": "scroll",
        "sort": "book_scroll",
        "components": "srg_meteorite_source/5,zettel/1",
        "trait": "BossScroll",
        "tag": "noShop",
    },
    "srg_scroll_convergence": {
        "category": "scroll",
        "sort": "book_scroll",
        "components": "srg_meteorite_source/2,meat/10",
        "trait": "ArchivistScroll",
        "tag": "noShop",
    },
    # ── Furniture: Tier 1 — Wanderer (always available) ──────────────────
    "srg_aurora_lamp": {
        "name_JP": "オーロラランプ",
        "name": "aurora lamp",
        "category": "light",
        "_idRenderData": "@obj_S",
        "tiles": 1552,
        "recipeKey": "*",
        "factory": "srg_codex",
        "components": "srg_meteorite_source/1,glass/2",
        "defMat": "glass",
        "value": 2000,
        "LV": 5,
        "chance": 0,
        "weight": 800,
        "trait": "Torch",
        "lightData": "candle_small",
        "tag": "tourism",
        "detail": "A softly glowing lamp that flickers with aurora-like light. Fragments of meteorite refract starlight through the glass.",
    },
    "srg_constellation_rug": {
        "name_JP": "星座のラグ",
        "name": "constellation rug",
        "category": "deco",
        "_idRenderData": "@obj flat",
        "tiles": 1552,
        "recipeKey": "*",
        "factory": "srg_codex",
        "components": "srg_meteorite_source/1,texture/3",
        "defMat": "wool",
        "value": 1800,
        "LV": 5,
        "chance": 0,
        "weight": 2000,
        "detail": "A woven rug depicting star patterns. The constellation lines seem to shimmer faintly.",
    },
    # ── Furniture: Tier 2 — Seeker (200 GP) ──────────────────────────────
    "srg_starfall_table": {
        "name_JP": "星降りのテーブル",
        "name": "starfall table",
        "category": "table",
        "_idRenderData": "@obj",
        "tiles": 1552,
        "recipeKey": "*",
        "factory": "srg_codex",
        "components": "srg_meteorite_source/2,plank/4",
        "defMat": "oak",
        "value": 3200,
        "LV": 10,
        "chance": 0,
        "weight": 6000,
        "detail": "A sturdy table inlaid with fragments of meteorite. The surface catches light in strange ways.",
    },
    "srg_lunar_armchair": {
        "name_JP": "月のアームチェア",
        "name": "lunar armchair",
        "category": "chair",
        "_idRenderData": "@obj",
        "tiles": 1552,
        "recipeKey": "*",
        "factory": "srg_codex",
        "components": "srg_meteorite_source/2,texture/3,stick/2",
        "defMat": "wool",
        "value": 3600,
        "LV": 12,
        "chance": 0,
        "weight": 4500,
        "trait": "Chair",
        "detail": "A plush armchair upholstered with star-patterned fabric. The crescent moon motif on the backrest glows faintly.",
    },
    "srg_celestial_globe": {
        "name_JP": "天球儀",
        "name": "celestial globe",
        "category": "deco",
        "_idRenderData": "@obj_S",
        "tiles": 1552,
        "recipeKey": "*",
        "factory": "srg_codex",
        "components": "srg_meteorite_source/2,glass/3",
        "defMat": "glass",
        "value": 4000,
        "LV": 10,
        "chance": 0,
        "weight": 3000,
        "tag": "tourism",
        "detail": "A glass sphere on a brass stand showing constellations. The star-lines glow with captured starlight.",
    },
    # ── Furniture: Tier 3 — Researcher (500 GP) ──────────────────────────
    "srg_zodiac_dresser": {
        "name_JP": "星座のドレッサー",
        "name": "zodiac dresser",
        "category": "container",
        "_idRenderData": "@obj",
        "tiles": 1552,
        "recipeKey": "*",
        "factory": "srg_codex",
        "components": "srg_meteorite_source/3,plank/6,bolt/2",
        "defMat": "oak",
        "value": 5000,
        "LV": 18,
        "chance": 0,
        "weight": 12000,
        "trait": "Container,5,4,crate",
        "roomName": "Storeroom,Stockroom",
        "detail": "A dresser carved with zodiac symbols. Each drawer is perfectly balanced for storing stargazing instruments.",
    },
    "srg_cosmic_mirror": {
        "name_JP": "宇宙の鏡",
        "name": "cosmic mirror",
        "category": "_furniture",
        "_idRenderData": "@obj",
        "tiles": 1552,
        "recipeKey": "*",
        "factory": "srg_codex",
        "components": "srg_meteorite_source/3,glass/6,plank/2",
        "defMat": "glass",
        "value": 5500,
        "LV": 15,
        "chance": 0,
        "weight": 7000,
        "trait": "Mirror",
        "detail": "A standing mirror framed in meteorite-laced silver. The reflection ripples with astral distortion.",
    },
    "srg_planisphere_cabinet": {
        "name_JP": "天球キャビネット",
        "name": "planisphere cabinet",
        "category": "container",
        "_tileType": "ObjBig",
        "_idRenderData": "@obj",
        "tiles": 1552,
        "recipeKey": "*",
        "factory": "srg_codex",
        "components": "srg_meteorite_source/3,plank/8,glass/2",
        "defMat": "oak",
        "value": 6000,
        "LV": 20,
        "chance": 0,
        "weight": 15000,
        "trait": "BookShelf",
        "detail": "A tall cabinet displaying star charts behind glass doors. The planisphere mechanism rotates slowly.",
    },
    # ── Furniture: Tier 4 — Cosmos-Addled (1500 GP) ──────────────────────
    "srg_stardust_bed": {
        "name_JP": "星塵のベッド",
        "name": "stardust bed",
        "category": "bed",
        "_idRenderData": "@obj flat",
        "tiles": 1552,
        "recipeKey": "*",
        "factory": "srg_codex",
        "components": "srg_meteorite_source/4,texture/6,plank/4",
        "defMat": "cotton",
        "value": 8000,
        "LV": 25,
        "chance": 0,
        "weight": 10000,
        "trait": "Bed,2",
        "roomName": "Bedroom",
        "detail": "A bed woven with stardust fibers. Sleepers report dreams of distant nebulae.",
    },
    "srg_astral_chandelier": {
        "name_JP": "星霊のシャンデリア",
        "name": "astral chandelier",
        "category": "mount",
        "_tileType": "ObjFloat",
        "_idRenderData": "@obj ceil",
        "tiles": 1552,
        "recipeKey": "*",
        "factory": "srg_codex",
        "components": "srg_meteorite_source/4,ingot/4,gem/2",
        "defMat": "gold",
        "value": 9000,
        "LV": 28,
        "chance": 0,
        "weight": 6000,
        "trait": "Torch",
        "lightData": "candle_small",
        "tag": "tourism",
        "detail": "A chandelier of crystallized starlight suspended from the ceiling. Its light dances like captured constellations.",
    },
    # ── Furniture: Tier 5 — Cosmos-Applied (3000 GP) ─────────────────────
    "srg_meteorite_statue": {
        "name_JP": "隕石の彫像",
        "name": "meteorite statue",
        "category": "deco",
        "_tileType": "ObjBig",
        "_idRenderData": "@obj tall",
        "tiles": 1552,
        "recipeKey": "*",
        "factory": "srg_codex",
        "components": "srg_meteorite_source/5,cutstone/8,ingot/4",
        "defMat": "granite",
        "value": 15000,
        "LV": 35,
        "chance": 0,
        "weight": 35000,
        "elements": "fireproof/1,acidproof/1",
        "tag": "tourism",
        "detail": "A towering statue carved from meteorite stone. Cosmic cracks glow with trapped starlight.",
    },
    # ── Furniture: Tier 6 — Understander (5000 GP) ───────────────────────
    "srg_nexus_core": {
        "name_JP": "巨大なる隕石の核",
        "name": "nexus meteor core",
        "category": "deco",
        "_tileType": "ObjBig",
        "_idRenderData": "@obj_LV",
        "tiles": 503,
        "defMat": "granite",
        "value": 50000,
        "LV": 50,
        "chance": 0,
        "weight": 99999,
        "elements": "fireproof/1,acidproof/1",
        "tag": "tourism,noShop",
        "detail": "A colossal, perfectly preserved meteor core pulled from an astral rift. It hums with the terrifying song of deep space.",
    },
    "srg_eclipse_hearth": {
        "name_JP": "蝕のかまど",
        "name": "eclipse hearth",
        "category": "crafter",
        "_idRenderData": "@obj",
        "tiles": 1552,
        "recipeKey": "*",
        "factory": "srg_codex",
        "components": "srg_meteorite_source/6,cutstone/10,brick/6,ingot/4",
        "defMat": "granite",
        "value": 20000,
        "LV": 40,
        "chance": 0,
        "weight": 25000,
        "trait": "Hearth,cooking",
        "lightData": "fireplace",
        "idExtra": "fireplace",
        "idToggleExtra": "fireplace",
        "idSound": "amb_fire",
        "roomName": "Kitchen,Dining Room",
        "detail": "A grand hearth carved with eclipse motifs. The flames burn with starfire and can cook any meal.",
    },
    # ── Interactive Furniture: Online Multiplayer Features ────────────────
    "srg_constellation_board": {
        "name_JP": "",
        "name": "constellation board",
        "category": "deco",
        "_tileType": "ObjBig",
        "_idRenderData": "@obj",
        "tiles": 1552,
        "defMat": "oak",
        "value": 3000,
        "LV": 15,
        "chance": 0,
        "quality": 3,
        "weight": 5000,
        "trait": "ConstellationBoard",
        "tag": "noShop,noWish",
        "roomName": "Observatory",
        "detail_JP": "",
        "detail": "A board displaying the five patron constellations of the season. Each tells a different cosmic story.",
    },
    "srg_geometry_orrery": {
        "name_JP": "",
        "name": "geometry orrery",
        "category": "deco",
        "_tileType": "ObjBig",
        "_idRenderData": "@obj",
        "tiles": 1552,
        "defMat": "bronze",
        "value": 4000,
        "LV": 15,
        "chance": 0,
        "quality": 3,
        "weight": 8000,
        "trait": "GeometryOrrery",
        "tag": "noShop,noWish",
        "roomName": "Observatory",
        "detail_JP": "",
        "detail": "A mechanical device tracking the dominant geometry of astral rifts across all Skyreaders.",
    },
    "srg_comet_heatmap_table": {
        "name_JP": "",
        "name": "astral contamination table",
        "category": "deco",
        "_tileType": "ObjBig",
        "_idRenderData": "@obj",
        "tiles": 1552,
        "defMat": "oak",
        "value": 2500,
        "LV": 10,
        "chance": 0,
        "quality": 3,
        "weight": 6000,
        "trait": "CometHeatmapTable",
        "tag": "noShop,noWish",
        "roomName": "Observatory",
        "detail_JP": "",
        "detail": "A table mapping astral contamination from meteor-touched townspeople and items.",
    },
    "srg_star_paper_shelf": {
        "name_JP": "",
        "name": "star paper shelf",
        "category": "deco",
        "_tileType": "ObjBig",
        "_idRenderData": "@obj",
        "tiles": 1552,
        "defMat": "oak",
        "value": 2000,
        "LV": 10,
        "chance": 0,
        "quality": 3,
        "weight": 7000,
        "trait": "StarPaperShelf",
        "tag": "noShop,noWish",
        "roomName": "Library",
        "detail_JP": "",
        "detail": "A shelf housing anonymous research notes submitted by Skyreaders across the world.",
    },
    "srg_star_paper_desk": {
        "name_JP": "",
        "name": "star paper writing desk",
        "category": "deco",
        "_idRenderData": "@obj",
        "tiles": 1552,
        "defMat": "oak",
        "value": 1800,
        "LV": 10,
        "chance": 0,
        "quality": 3,
        "weight": 4000,
        "trait": "StarPaperWritingDesk",
        "tag": "noShop,noWish",
        "roomName": "Library",
        "detail_JP": "",
        "detail": "A writing desk for composing star papers. The ink shimmers faintly.",
    },
}


EXPECTED_CHARAS = {
    "srg_growth": {
        "_id": 900101,
        "name_JP": "不完全なイースの成長体",
        "name": "Incomplete Yith Growth",
        "aka_JP": "",
        "aka": "",
        "idActor": "",
        "sort": "",
        "size": "",
        "_idRenderData": "@chara_L",
        "tiles": 16,
        "tiles_snow": "",
        "colorMod": "",
        "components": "",
        "defMat": "",
        "LV": 18,
        "chance": 0,
        "quality": "",
        "hostility": "Enemy",
        "biome": "",
        "tag": "noRandomProduct",
        "trait": "",
        "race": "yith",
        "job": "predator",
        "tactics": "",
        "aiIdle": "",
        "aiParam": "2,40,0",
        "actCombat": "Wait,SpSilence/5,ActGazeMutation/10,hand_Mind/50,SpSummonMonster/5",
        "mainElement": "",
        "elements": "resNether/10",
        "equip": "",
        "loot": "",
        "category": "",
        "filter": "",
        "gachaFilter": "",
        "tone": "",
        "actIdle": "",
        "lightData": "",
        "idExtra": "",
        "bio": "",
        "faith": "",
        "works": "",
        "hobbies": "",
        "idText": "",
        "moveAnime": "",
        "factory": "",
        "components_2": "",
        "recruitItems": "",
        "detail_JP": "",
        "detail": "",
    },
    "srg_arkyn": {
        "_id": 900102,
        "name_JP": "",
        "name": "Arkyn, Keeper of Stars",
        "aka_JP": "",
        "aka": "",
        "idActor": "",
        "sort": "",
        "size": "",
        "_idRenderData": "@chara",
        "tiles": 806,
        "tiles_snow": "",
        "colorMod": "",
        "components": "",
        "defMat": "",
        "LV": 50,
        "chance": 0,
        "quality": 4,
        "hostility": "Friend",
        "biome": "",
        "tag": "neutral,noRandomProduct",
        "trait": "UniqueChara",
        "race": "elea",
        "job": "wizard",
        "tactics": "",
        "aiIdle": "",
        "aiParam": "",
        "actCombat": "",
        "mainElement": "",
        "elements": "",
        "equip": "",
        "loot": "",
        "category": "",
        "filter": "",
        "gachaFilter": "",
        "tone": "",
        "actIdle": "",
        "lightData": "",
        "idExtra": "",
        "bio": "",
        "faith": "",
        "works": "Research",
        "hobbies": "Read",
        "idText": "",
        "moveAnime": "",
        "factory": "",
        "components_2": "",
        "recruitItems": "",
        "detail_JP": "",
        "detail": "",
    },
    "srg_umbryon": {
        "_id": 900501,
        "name_JP": "",
        "name": "Umbryon, Herald of Eternal Rot",
        "aka_JP": "",
        "aka": "",
        "idActor": "",
        "sort": "",
        "size": "",
        "_idRenderData": "@chara",
        "tiles": 1502,
        "tiles_snow": "",
        "colorMod": "",
        "components": "",
        "defMat": "",
        "LV": 35,
        "chance": 0,
        "quality": 3,
        "hostility": "Enemy",
        "biome": "",
        "tag": "boss,noRandomProduct",
        "trait": "UniqueMonster",
        "race": "lich",
        "job": "warrior",
        "tactics": "",
        "aiIdle": "",
        "aiParam": "",
        "actCombat": "bolt_Darkness/50,miasma_Poison/30,ActDrainBlood/60",
        "mainElement": "Darkness",
        "elements": "life/100,resDarkness/20,END/10,WIL/10",
        "equip": "",
        "loot": "",
        "category": "",
        "filter": "",
        "gachaFilter": "",
        "tone": "",
        "actIdle": "",
        "lightData": "",
        "idExtra": "",
        "bio": "",
        "faith": "",
        "works": "",
        "hobbies": "",
        "idText": "",
        "moveAnime": "",
        "factory": "",
        "components_2": "",
        "recruitItems": "",
        "detail_JP": "",
        "detail": "",
    },
    "srg_solaris": {
        "_id": 900502,
        "name_JP": "",
        "name": "Solaris, Inferno of the Fallen Star",
        "aka_JP": "",
        "aka": "",
        "idActor": "",
        "sort": "",
        "size": "",
        "_idRenderData": "@chara",
        "tiles": 1713,
        "tiles_snow": "",
        "colorMod": "",
        "components": "",
        "defMat": "",
        "LV": 40,
        "chance": 0,
        "quality": 3,
        "hostility": "Enemy",
        "biome": "",
        "tag": "boss,noRandomProduct",
        "trait": "UniqueMonster",
        "race": "spirit",
        "job": "warmage",
        "tactics": "",
        "aiIdle": "",
        "aiParam": "",
        "actCombat": "breathe_Fire/70,SpMeteor/10,bolt_Fire/50",
        "mainElement": "Fire",
        "elements": "resFire/30,life/80,MAG/15,mana/30",
        "equip": "",
        "loot": "",
        "category": "",
        "filter": "",
        "gachaFilter": "",
        "tone": "",
        "actIdle": "",
        "lightData": "",
        "idExtra": "",
        "bio": "",
        "faith": "",
        "works": "",
        "hobbies": "",
        "idText": "",
        "moveAnime": "",
        "factory": "",
        "components_2": "",
        "recruitItems": "",
        "detail_JP": "",
        "detail": "",
    },
    "srg_erevor": {
        "_id": 900503,
        "name_JP": "",
        "name": "Erevor, The Abyssal Maw",
        "aka_JP": "",
        "aka": "",
        "idActor": "",
        "sort": "",
        "size": "",
        "_idRenderData": "@chara_L",
        "tiles": 104,
        "tiles_snow": "",
        "colorMod": "",
        "components": "",
        "defMat": "",
        "LV": 45,
        "chance": 0,
        "quality": 3,
        "hostility": "Enemy",
        "biome": "",
        "tag": "boss,noRandomProduct",
        "trait": "UniqueMonster",
        "race": "dragon",
        "job": "predator",
        "tactics": "",
        "aiIdle": "",
        "aiParam": "",
        "actCombat": "SpGravity/50,ActRush,breathe_Void/30",
        "mainElement": "Impact",
        "elements": "life/100,PDR/30,STR/15,END/15",
        "equip": "",
        "loot": "",
        "category": "",
        "filter": "",
        "gachaFilter": "",
        "tone": "",
        "actIdle": "",
        "lightData": "",
        "idExtra": "",
        "bio": "",
        "faith": "",
        "works": "",
        "hobbies": "",
        "idText": "",
        "moveAnime": "",
        "factory": "",
        "components_2": "",
        "recruitItems": "",
        "detail_JP": "",
        "detail": "",
    },
    "srg_quasarix": {
        "_id": 900504,
        "name_JP": "",
        "name": "Quasarix, Devourer of Light",
        "aka_JP": "",
        "aka": "",
        "idActor": "",
        "sort": "",
        "size": "",
        "_idRenderData": "@chara",
        "tiles": 2317,
        "tiles_snow": "",
        "colorMod": "",
        "components": "",
        "defMat": "",
        "LV": 50,
        "chance": 0,
        "quality": 3,
        "hostility": "Enemy",
        "biome": "",
        "tag": "boss,noRandomProduct",
        "trait": "UniqueMonster",
        "race": "god",
        "job": "gunner",
        "tactics": "",
        "aiIdle": "",
        "aiParam": "",
        "actCombat": "SpSilence/50,ActGazeDim/30,bolt_Magic/60,SpBane/20",
        "mainElement": "Darkness",
        "elements": "resMagic/20,resDarkness/20,life/80,mana/50,MAG/15",
        "equip": "",
        "loot": "",
        "category": "",
        "filter": "",
        "gachaFilter": "",
        "tone": "",
        "actIdle": "",
        "lightData": "",
        "idExtra": "",
        "bio": "",
        "faith": "",
        "works": "",
        "hobbies": "",
        "idText": "",
        "moveAnime": "",
        "factory": "",
        "components_2": "",
        "recruitItems": "",
        "detail_JP": "",
        "detail": "",
    },
    "srg_archivist": {
        "_id": 900505,
        "name_JP": "",
        "name": "Astral Archivist",
        "aka_JP": "",
        "aka": "",
        "idActor": "",
        "sort": "",
        "size": "",
        "_idRenderData": "@chara",
        "tiles": 1216,
        "tiles_snow": "",
        "colorMod": "",
        "components": "",
        "defMat": "",
        "LV": 30,
        "chance": 0,
        "quality": 4,
        "hostility": "Neutral",
        "biome": "",
        "tag": "neutral,noRandomProduct",
        "trait": "Archivist",
        "race": "elea",
        "job": "pianist",
        "tactics": "",
        "aiIdle": "",
        "aiParam": "",
        "actCombat": "SpHealHeavy/60,SpHero/50/pt",
        "mainElement": "",
        "elements": "featHealer/1,reading/10,MAG/10,WIL/10",
        "equip": "",
        "loot": "",
        "category": "",
        "filter": "",
        "gachaFilter": "",
        "tone": "",
        "actIdle": "",
        "lightData": "",
        "idExtra": "",
        "bio": "",
        "faith": "",
        "works": "",
        "hobbies": "",
        "idText": "",
        "moveAnime": "",
        "factory": "",
        "components_2": "",
        "recruitItems": "",
        "detail_JP": "",
        "detail": "",
    },
    "srg_yith_hound": {
        "_id": 900601,
        "name_JP": "",
        "name": "Stray Yith Hound",
        "aka_JP": "",
        "aka": "",
        "idActor": "",
        "sort": "",
        "size": "",
        "_idRenderData": "@chara",
        "tiles": 2823,
        "tiles_snow": "",
        "colorMod": "",
        "components": "",
        "defMat": "",
        "LV": 10,
        "chance": 0,
        "quality": "",
        "hostility": "Enemy",
        "biome": "",
        "tag": "noRandomProduct",
        "trait": "",
        "race": "yith",
        "job": "predator",
        "tactics": "",
        "aiIdle": "",
        "aiParam": "2,40,0",
        "actCombat": "ActEntangle/40,hand_Mind/50",
        "mainElement": "",
        "elements": "resNether/5",
        "equip": "",
        "loot": "",
        "category": "",
        "filter": "",
        "gachaFilter": "",
        "tone": "",
        "actIdle": "",
        "lightData": "",
        "idExtra": "",
        "bio": "",
        "faith": "",
        "works": "",
        "hobbies": "",
        "idText": "",
        "moveAnime": "",
        "factory": "",
        "components_2": "",
        "recruitItems": "",
        "detail_JP": "",
        "detail": "",
    },
    "srg_yith_drone": {
        "_id": 900602,
        "name_JP": "",
        "name": "Astral Yith Drone",
        "aka_JP": "",
        "aka": "",
        "idActor": "",
        "sort": "",
        "size": "",
        "_idRenderData": "@chara",
        "tiles": 1627,
        "tiles_snow": "",
        "colorMod": "",
        "components": "",
        "defMat": "",
        "LV": 25,
        "chance": 0,
        "quality": "",
        "hostility": "Enemy",
        "biome": "",
        "tag": "noRandomProduct",
        "trait": "",
        "race": "yith",
        "job": "predator",
        "tactics": "",
        "aiIdle": "",
        "aiParam": "2,40,0",
        "actCombat": "ActEntangle/50,SpSilence/10,hand_Mind/50",
        "mainElement": "Nerve",
        "elements": "resNether/10,resMind/5",
        "equip": "",
        "loot": "",
        "category": "",
        "filter": "",
        "gachaFilter": "",
        "tone": "",
        "actIdle": "",
        "lightData": "",
        "idExtra": "",
        "bio": "",
        "faith": "",
        "works": "",
        "hobbies": "",
        "idText": "",
        "moveAnime": "",
        "factory": "",
        "components_2": "",
        "recruitItems": "",
        "detail_JP": "",
        "detail": "",
    },
    "srg_yith_weaver": {
        "_id": 900603,
        "name_JP": "",
        "name": "Cosmic Yith Weaver",
        "aka_JP": "",
        "aka": "",
        "idActor": "",
        "sort": "",
        "size": "",
        "_idRenderData": "@chara_L",
        "tiles": 16,
        "tiles_snow": "",
        "colorMod": "",
        "components": "",
        "defMat": "",
        "LV": 40,
        "chance": 0,
        "quality": "",
        "hostility": "Enemy",
        "biome": "",
        "tag": "noRandomProduct",
        "trait": "",
        "race": "yith",
        "job": "warmage",
        "tactics": "",
        "aiIdle": "",
        "aiParam": "2,40,0",
        "actCombat": "SpGravity/40,ActGazeMutation/15,ball_Chaos/30,hand_Mind/50",
        "mainElement": "Chaos",
        "elements": "resChaos/10,resNether/10,MAG/10",
        "equip": "",
        "loot": "",
        "category": "",
        "filter": "",
        "gachaFilter": "",
        "tone": "",
        "actIdle": "",
        "lightData": "",
        "idExtra": "",
        "bio": "",
        "faith": "",
        "works": "",
        "hobbies": "",
        "idText": "",
        "moveAnime": "",
        "factory": "",
        "components_2": "",
        "recruitItems": "",
        "detail_JP": "",
        "detail": "",
    },
    "srg_yith_ancient": {
        "_id": 900604,
        "name_JP": "",
        "name": "Void-Touched Yith Ancient",
        "aka_JP": "",
        "aka": "",
        "idActor": "",
        "sort": "",
        "size": "",
        "_idRenderData": "@chara_L",
        "tiles": 15,
        "tiles_snow": "",
        "colorMod": "",
        "components": "",
        "defMat": "",
        "LV": 60,
        "chance": 0,
        "quality": "",
        "hostility": "Enemy",
        "biome": "",
        "tag": "noRandomProduct",
        "trait": "",
        "race": "yith",
        "job": "executioner",
        "tactics": "",
        "aiIdle": "",
        "aiParam": "2,40,0",
        "actCombat": "ActDrainBlood/60,ActGazeInsane/30,SpBane/20,ActEntangle/40",
        "mainElement": "Darkness",
        "elements": "resNether/15,resDarkness/10,life/50,STR/10,END/10",
        "equip": "",
        "loot": "",
        "category": "",
        "filter": "",
        "gachaFilter": "",
        "tone": "",
        "actIdle": "",
        "lightData": "",
        "idExtra": "",
        "bio": "",
        "faith": "",
        "works": "",
        "hobbies": "",
        "idText": "",
        "moveAnime": "",
        "factory": "",
        "components_2": "",
        "recruitItems": "",
        "detail_JP": "",
        "detail": "",
    },
    "srg_yith_behemoth": {
        "_id": 900605,
        "name_JP": "",
        "name": "Eldritch Yith Behemoth",
        "aka_JP": "",
        "aka": "",
        "idActor": "",
        "sort": "",
        "size": "",
        "_idRenderData": "@chara_L",
        "tiles": 110,
        "tiles_snow": "",
        "colorMod": "",
        "components": "",
        "defMat": "",
        "LV": 85,
        "chance": 0,
        "quality": "",
        "hostility": "Enemy",
        "biome": "",
        "tag": "noRandomProduct",
        "trait": "",
        "race": "yith",
        "job": "predator",
        "tactics": "",
        "aiIdle": "",
        "aiParam": "2,40,0",
        "actCombat": "ActDrainBlood/80,ActGazeInsane/40,SpBane/30,SpSummonMonster/15,breathe_Chaos/20",
        "mainElement": "Chaos",
        "elements": "resChaos/15,resNether/15,resMind/10,life/100,mana/50,STR/15,END/15,MAG/10",
        "equip": "",
        "loot": "",
        "category": "",
        "filter": "",
        "gachaFilter": "",
        "tone": "",
        "actIdle": "",
        "lightData": "",
        "idExtra": "",
        "bio": "",
        "faith": "",
        "works": "",
        "hobbies": "",
        "idText": "",
        "moveAnime": "",
        "factory": "",
        "components_2": "",
        "recruitItems": "",
        "detail_JP": "",
        "detail": "",
    },
}


def find_rows(ws, col_map):
    rows = {}
    for row in range(4, ws.max_row + 1):
        item_id = ws.cell(row=row, column=col_map["id"]).value
        if isinstance(item_id, str):
            rows[item_id] = row
    return rows


def set_field(ws, row, col_map, key, value):
    col = col_map[key]
    cell = ws.cell(row=row, column=col)
    if value == "":
        value = None
    if cell.value != value:
        print(f"  FIX: {ws.cell(row=row, column=1).value} {key}: {cell.value!r} -> {value!r}")
        cell.value = value
        return 1
    return 0


def normalize_chara_typos(ws):
    rows = find_rows(ws, CHARA_COL)
    typo = rows.get("srg_ervor")
    correct = rows.get("srg_erevor")
    if typo is not None and correct is None:
        ws.cell(row=typo, column=CHARA_COL["id"]).value = "srg_erevor"
        print(f"  FIX: srg_ervor id: 'srg_ervor' -> 'srg_erevor' at row {typo}")
        return 1
    if typo is not None and correct is not None:
        ws.delete_rows(typo, 1)
        print(f"  DELETE: duplicate typo row srg_ervor at row {typo}")
        return 1
    return 0


def delete_forbidden_chara_rows(ws):
    rows = find_rows(ws, CHARA_COL)
    deleted = 0
    for chara_id in sorted(FORBIDDEN_CHARA_IDS):
        row = rows.get(chara_id)
        if row is not None:
            ws.delete_rows(row, 1)
            deleted += 1
            print(f"  DELETE: forbidden test Chara row {chara_id} at row {row}")
            rows = find_rows(ws, CHARA_COL)
    return deleted


def ensure_thing_rows(ws):
    rows = find_rows(ws, THING_COL)
    changed = 0
    added = 0

    for item_id, fields in EXPECTED_THINGS.items():
        row = rows.get(item_id)
        if row is None:
            row = ws.max_row + 1
            ws.cell(row=row, column=THING_COL["id"], value=item_id)
            rows[item_id] = row
            added += 1
            print(f"  ADD: {item_id} at row {row}")

        for key, value in fields.items():
            changed += set_field(ws, row, THING_COL, key, value)

        sort_value = ws.cell(row=row, column=THING_COL["sort_value"]).value
        if sort_value is not None and not isinstance(sort_value, int):
            print(f"  CLEAN: {item_id} numeric sort column {sort_value!r} -> None")
            ws.cell(row=row, column=THING_COL["sort_value"]).value = None
            changed += 1

    return added, changed


def ensure_chara_rows(ws):
    changed = delete_forbidden_chara_rows(ws)
    changed += normalize_chara_typos(ws)
    rows = find_rows(ws, CHARA_COL)
    added = 0

    for chara_id, fields in EXPECTED_CHARAS.items():
        row = rows.get(chara_id)
        if row is None:
            row = ws.max_row + 1
            ws.cell(row=row, column=CHARA_COL["id"], value=chara_id)
            rows[chara_id] = row
            added += 1
            print(f"  ADD: {chara_id} at row {row}")

        for key, value in fields.items():
            changed += set_field(ws, row, CHARA_COL, key, value)

    return added, changed


def main():
    wb = openpyxl.load_workbook(SOURCE_CARD)
    ws_thing = wb["Thing"]
    ws_chara = wb["Chara"]

    thing_added, thing_changed = ensure_thing_rows(ws_thing)
    chara_added, chara_changed = ensure_chara_rows(ws_chara)
    added = thing_added + chara_added
    changed = thing_changed + chara_changed

    if added or changed:
        wb.save(SOURCE_CARD)
        print(f"Saved SourceCard.xlsx ({added} added, {changed} fixed).")
    else:
        print("SourceCard.xlsx already matches expected Skyreader rows.")

    wb.close()
    if added or changed:
        normalize_shared_strings(SOURCE_CARD)
        print("Normalized shared strings in", SOURCE_CARD)

    # Deploy textures and SourceCard if they exist
    root_dir = Path(__file__).resolve().parent.parent.parent
    texture_src = root_dir / "Texture"
    target_package_dir = Path(r"D:\Steam\steamapps\common\Elin\Package\SkyreaderGuild")
    texture_dest = target_package_dir / "Texture"
    
    try:
        # Deploy Texture dir
        if texture_src.exists():
            if texture_dest.exists():
                shutil.rmtree(texture_dest)
            shutil.copytree(texture_src, texture_dest)
            print(f"Deployed {texture_src.name} to {texture_dest}")
        else:
            print(f"No {texture_src.name} directory found, skipping texture deployment.")

        # Deploy SourceCard
        source_card_src = root_dir / "LangMod" / "EN" / "SourceCard.xlsx"
        source_card_dest_dir = target_package_dir / "LangMod" / "EN"
        if source_card_src.exists():
            source_card_dest_dir.mkdir(parents=True, exist_ok=True)
            shutil.copy2(source_card_src, source_card_dest_dir / "SourceCard.xlsx")
            print(f"Deployed SourceCard.xlsx to {source_card_dest_dir}")
        else:
            print(f"No SourceCard.xlsx found at {source_card_src}, skipping.")
    except PermissionError as exc:
        print(f"Deployment skipped: {exc}")

if __name__ == "__main__":
    main()

