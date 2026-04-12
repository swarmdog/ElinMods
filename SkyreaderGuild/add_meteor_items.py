"""Ensure Skyreader Guild source rows are present and importer-safe.

Elin's NPOI-based importer expects workbook strings through sharedStrings.xml.
openpyxl writes inlineStr cells by default, which Elin imports as empty values.
This script validates the mod's key Thing rows, fixes known data drift, and
normalizes workbook strings after saving.
"""

import io
import shutil
import sys
import tempfile
import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path

import openpyxl

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

ROOT = Path(__file__).resolve().parent
SOURCE_CARD = ROOT / "LangMod" / "EN" / "SourceCard.xlsx"

NS_MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
NS_REL = "http://schemas.openxmlformats.org/package/2006/relationships"
NS_CT = "http://schemas.openxmlformats.org/package/2006/content-types"
NS_DOC_REL = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"

ET.register_namespace("", NS_MAIN)
ET.register_namespace("r", NS_DOC_REL)


def read_shared_strings(zin):
    if "xl/sharedStrings.xml" not in zin.namelist():
        return []
    root = ET.fromstring(zin.read("xl/sharedStrings.xml"))
    values = []
    for si in root.findall(f"{{{NS_MAIN}}}si"):
        parts = []
        for t in si.iter(f"{{{NS_MAIN}}}t"):
            parts.append(t.text or "")
        values.append("".join(parts))
    return values


def normalize_shared_strings(path):
    """Convert all string cells to shared strings for Elin's source importer."""
    shared = []
    shared_index = {}

    def intern(text):
        if text not in shared_index:
            shared_index[text] = len(shared)
            shared.append(text)
        return shared_index[text]

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        tmp_path = Path(tmp.name)

    try:
        with zipfile.ZipFile(path, "r") as zin:
            old_shared = read_shared_strings(zin)
            with zipfile.ZipFile(tmp_path, "w", zipfile.ZIP_DEFLATED) as zout:
                for info in zin.infolist():
                    if info.filename == "xl/sharedStrings.xml":
                        continue

                    data = zin.read(info.filename)
                    if info.filename.startswith("xl/worksheets/sheet") and info.filename.endswith(".xml"):
                        root = ET.fromstring(data)
                        for cell in root.iter(f"{{{NS_MAIN}}}c"):
                            cell_type = cell.get("t")
                            if cell_type not in ("inlineStr", "s", "str"):
                                continue

                            text = ""
                            if cell_type == "inlineStr":
                                inline = cell.find(f"{{{NS_MAIN}}}is")
                                if inline is not None:
                                    text = "".join(t.text or "" for t in inline.iter(f"{{{NS_MAIN}}}t"))
                            else:
                                v = cell.find(f"{{{NS_MAIN}}}v")
                                if v is not None:
                                    if cell_type == "s":
                                        idx = int(v.text or "0")
                                        text = old_shared[idx] if 0 <= idx < len(old_shared) else ""
                                    else:
                                        text = v.text or ""

                            for child in list(cell):
                                cell.remove(child)
                            cell.set("t", "s")
                            v = ET.SubElement(cell, f"{{{NS_MAIN}}}v")
                            v.text = str(intern(text))
                        data = ET.tostring(root, encoding="utf-8", xml_declaration=True)

                    elif info.filename == "[Content_Types].xml":
                        root = ET.fromstring(data)
                        part_name = "/xl/sharedStrings.xml"
                        exists = any(
                            child.tag == f"{{{NS_CT}}}Override" and child.get("PartName") == part_name
                            for child in root
                        )
                        if not exists:
                            ET.SubElement(
                                root,
                                f"{{{NS_CT}}}Override",
                                {
                                    "PartName": part_name,
                                    "ContentType": (
                                        "application/vnd.openxmlformats-officedocument."
                                        "spreadsheetml.sharedStrings+xml"
                                    ),
                                },
                            )
                        data = ET.tostring(root, encoding="utf-8", xml_declaration=True)

                    elif info.filename == "xl/_rels/workbook.xml.rels":
                        root = ET.fromstring(data)
                        exists = any(
                            child.tag == f"{{{NS_REL}}}Relationship"
                            and child.get("Type") == f"{NS_DOC_REL}/sharedStrings"
                            for child in root
                        )
                        if not exists:
                            used = {
                                int((child.get("Id") or "rId0").replace("rId", ""))
                                for child in root
                                if (child.get("Id") or "").startswith("rId")
                            }
                            next_id = 1
                            while next_id in used:
                                next_id += 1
                            ET.SubElement(
                                root,
                                f"{{{NS_REL}}}Relationship",
                                {
                                    "Id": f"rId{next_id}",
                                    "Type": f"{NS_DOC_REL}/sharedStrings",
                                    "Target": "sharedStrings.xml",
                                },
                            )
                        data = ET.tostring(root, encoding="utf-8", xml_declaration=True)

                    zout.writestr(info, data)

                sst = ET.Element(
                    f"{{{NS_MAIN}}}sst",
                    {"count": str(len(shared)), "uniqueCount": str(len(shared))},
                )
                for text in shared:
                    si = ET.SubElement(sst, f"{{{NS_MAIN}}}si")
                    t = ET.SubElement(si, f"{{{NS_MAIN}}}t")
                    if text != text.strip():
                        t.set("{http://www.w3.org/XML/1998/namespace}space", "preserve")
                    t.text = text
                zout.writestr(
                    "xl/sharedStrings.xml",
                    ET.tostring(sst, encoding="utf-8", xml_declaration=True),
                )

        shutil.move(tmp_path, path)
    finally:
        if tmp_path.exists():
            tmp_path.unlink()


THING_COL = {
    "id": 1,
    "name_JP": 2,
    "unit_JP": 4,
    "name": 6,
    "category": 9,
    "sort": 10,
    "sort_value": 11,
    "_tileType": 12,
    "_idRenderData": 13,
    "tiles": 14,
    "recipeKey": 21,
    "factory": 22,
    "components": 23,
    "defMat": 25,
    "value": 27,
    "LV": 28,
    "chance": 29,
    "HP": 31,
    "weight": 32,
    "trait": 34,
    "tag": 46,
    "detail_JP": 51,
    "detail": 52,
}


CHARA_COL = {
    "id": 1,
    "_id": 2,
    "name_JP": 3,
    "name": 4,
    "aka_JP": 5,
    "aka": 6,
    "idActor": 7,
    "sort": 8,
    "size": 9,
    "_idRenderData": 10,
    "tiles": 11,
    "tiles_snow": 12,
    "colorMod": 13,
    "components": 14,
    "defMat": 15,
    "LV": 16,
    "chance": 17,
    "quality": 18,
    "hostility": 19,
    "biome": 20,
    "tag": 21,
    "trait": 22,
    "race": 23,
    "job": 24,
    "tactics": 25,
    "aiIdle": 26,
    "aiParam": 27,
    "actCombat": 28,
    "mainElement": 29,
    "elements": 30,
    "equip": 31,
    "loot": 32,
    "category": 33,
    "filter": 34,
    "gachaFilter": 35,
    "tone": 36,
    "actIdle": 37,
    "lightData": 38,
    "idExtra": 39,
    "bio": 40,
    "faith": 41,
    "works": 42,
    "hobbies": 43,
    "idText": 44,
    "moveAnime": 45,
    "factory": 46,
    "components_2": 47,
    "recruitItems": 48,
    "detail_JP": 49,
    "detail": 50,
}


FORBIDDEN_CHARA_IDS = {
    "chickchicken",
    "chickchick",
    "chickchik",
    "test_chicken",
}


EXPECTED_THINGS = {
    "srg_starchart": {
        "name_JP": "淡い星図",
        "unit_JP": "冊",
        "name": "faint starchart",
        "category": "map",
        "sort": "book",
        "_idRenderData": "obj_S flat",
        "tiles": 1611,
        "components": "texture",
        "defMat": "paper",
        "value": 100,
        "LV": 1,
        "weight": 150,
        "trait": "ScrollMap",
        "detail": "The text on the page seems to waver faintly...",
    },
    "srg_codex": {
        "name_JP": "占星術写本",
        "name": "astrological codex",
        "category": "crafter",
        "_tileType": "ObjBig",
        "_idRenderData": "obj",
        "tiles": 1552,
        "recipeKey": "*",
        "factory": "workbench",
        "components": "srg_meteorite_source/4,ore_gem/12,log/4,ingot/9",
        "defMat": "paper",
        "value": 5000,
        "LV": 10,
        "chance": 0,
        "HP": 80,
        "weight": 12000,
        "trait": "AstrologicalCodex,reading",
        "detail_JP": "A codex and instruments of strange design.",
        "detail": "A codex and instruments of strange design.",
    },
    "srg_astral_extractor": {
        "name_JP": "星霊抽出器",
        "name": "astral extractor",
        "category": "_item",
        "_idRenderData": "obj_S flat",
        "tiles": 1208,
        "recipeKey": "*",
        "factory": "srg_codex",
        "components": "srg_meteorite_source/1,potion_empty/1",
        "defMat": "glass",
        "value": 300,
        "LV": 1,
        "chance": 0,
        "weight": 200,
        "trait": "AstralExtractor",
        "detail_JP": "A device that pulls stray starlight from meteor-touched beings and objects.",
        "detail": "A device that pulls stray starlight from meteor-touched beings and objects.",
    },
    "srg_meteor_core": {
        "name_JP": "隕石の核",
        "name": "meteor core",
        "category": "junk",
        "sort": "junk",
        "_idRenderData": "obj_S",
        "tiles": 503,
        "defMat": "granite",
        "value": 500,
        "LV": 1,
        "weight": 5000,
        "trait": "MeteorCore",
        "detail": "A pulsing core of extraterrestrial origin. It thrums with energy.",
    },
    "srg_meteorite_source": {
        "name_JP": "隕石素材",
        "name": "meteorite source",
        "category": "ore",
        "sort": "resource_ore",
        "_idRenderData": "obj_S flat",
        "tiles": 530,
        "defMat": "gold",
        "value": 200,
        "LV": 1,
        "weight": 500,
        "trait": "ResourceMain",
        "detail": "A fragment of fallen star, rich with rare minerals.",
    },
    "srg_debris": {
        "name_JP": "衝突破片",
        "name": "impact debris",
        "category": "junk",
        "sort": "junk",
        "_idRenderData": "obj_S",
        "tiles": 503,
        "defMat": "granite",
        "value": 10,
        "LV": 1,
        "weight": 2000,
        "detail": "Scorched fragments from a meteor impact.",
    },
    "srg_weave_stars": {
        "name_JP": "星の織物",
        "name": "weave of the stars",
        "category": "_item",
        "_idRenderData": "obj_S flat",
        "tiles": 504,
        "recipeKey": "*",
        "factory": "srg_codex",
        "components": "srg_meteorite_source/2",
        "defMat": "gold",
        "value": 1000,
        "LV": 1,
        "chance": 0,
        "weight": 100,
        "trait": "StarImbuement,weave",
        "detail_JP": "Starlight folded into a form that can settle into armor.",
        "detail": "Starlight folded into a form that can settle into armor.",
    },
    "srg_starforge": {
        "name_JP": "星鍛の火花",
        "name": "starforge spark",
        "category": "_item",
        "_idRenderData": "obj_S flat",
        "tiles": 530,
        "recipeKey": "*",
        "factory": "srg_codex",
        "components": "srg_meteorite_source/2",
        "defMat": "gold",
        "value": 1200,
        "LV": 1,
        "chance": 0,
        "weight": 100,
        "trait": "StarImbuement,forge",
        "detail_JP": "A shard of intent for awakening weapons beneath starlight.",
        "detail": "A shard of intent for awakening weapons beneath starlight.",
    },
    "srg_astral_portal": {
        "name_JP": "星霊の門",
        "name": "astral portal",
        "category": "_item",
        "_idRenderData": "obj_S",
        "tiles": 1196,
        "defMat": "glass",
        "value": 0,
        "LV": 1,
        "weight": 99999,
        "trait": "AstralPortal",
        "tag": "noShop,noWish",
        "detail_JP": "A shimmering gateway to an astral rift.",
        "detail": "A shimmering gateway to an astral rift.",
    },
    "srg_scroll_twilight": {
        "category": "scroll",
        "sort": "book_scroll",
        "components": "srg_meteorite_source/5,zettel/1",
        "trait": "BossScroll",
        "tag": "noShop",
    },
    "srg_scroll_radiance": {
        "category": "scroll",
        "sort": "book_scroll",
        "components": "srg_meteorite_source/5,zettel/1",
        "trait": "BossScroll",
        "tag": "noShop",
    },
    "srg_scroll_abyss": {
        "category": "scroll",
        "sort": "book_scroll",
        "components": "srg_meteorite_source/5,zettel/1",
        "trait": "BossScroll",
        "tag": "noShop",
    },
    "srg_scroll_nova": {
        "category": "scroll",
        "sort": "book_scroll",
        "components": "srg_meteorite_source/5,zettel/1",
        "trait": "BossScroll",
        "tag": "noShop",
    },
    "srg_scroll_convergence": {
        "category": "scroll",
        "sort": "book_scroll",
        "components": "srg_meteorite_source/2,meat/10",
        "trait": "ArchivistScroll",
        "tag": "noShop",
    },
}


EXPECTED_CHARAS = {
    "srg_growth": {
        "_id": 900101,
        "name_JP": "不完全なイースの成長体",
        "name": "Incomplete Yith Growth",
        "aka_JP": "",
        "aka": "",
        "idActor": "",
        "sort": "",
        "size": "",
        "_idRenderData": "chara_L",
        "tiles": 16,
        "tiles_snow": "",
        "colorMod": "",
        "components": "",
        "defMat": "",
        "LV": 18,
        "chance": 0,
        "quality": "",
        "hostility": "Enemy",
        "biome": "",
        "tag": "noRandomProduct",
        "trait": "",
        "race": "yith",
        "job": "predator",
        "tactics": "",
        "aiIdle": "",
        "aiParam": "2,40,0",
        "actCombat": "Wait,SpSilence/5,ActGazeMutation/10,hand_Mind/50,SpSummonMonster/5",
        "mainElement": "",
        "elements": "resNether/10",
        "equip": "",
        "loot": "",
        "category": "",
        "filter": "",
        "gachaFilter": "",
        "tone": "",
        "actIdle": "",
        "lightData": "",
        "idExtra": "",
        "bio": "",
        "faith": "",
        "works": "",
        "hobbies": "",
        "idText": "",
        "moveAnime": "",
        "factory": "",
        "components_2": "",
        "recruitItems": "",
        "detail_JP": "",
        "detail": "",
    },
    "srg_arkyn": {
        "_id": 900102,
        "name_JP": "",
        "name": "Arkyn, Keeper of Stars",
        "aka_JP": "",
        "aka": "",
        "idActor": "",
        "sort": "",
        "size": "",
        "_idRenderData": "",
        "tiles": 806,
        "tiles_snow": "",
        "colorMod": "",
        "components": "",
        "defMat": "",
        "LV": 50,
        "chance": 0,
        "quality": 4,
        "hostility": "Friend",
        "biome": "",
        "tag": "neutral,noRandomProduct",
        "trait": "UniqueChara",
        "race": "elea",
        "job": "wizard",
        "tactics": "",
        "aiIdle": "",
        "aiParam": "",
        "actCombat": "",
        "mainElement": "",
        "elements": "",
        "equip": "",
        "loot": "",
        "category": "",
        "filter": "",
        "gachaFilter": "",
        "tone": "",
        "actIdle": "",
        "lightData": "",
        "idExtra": "",
        "bio": "",
        "faith": "",
        "works": "Research",
        "hobbies": "Read",
        "idText": "",
        "moveAnime": "",
        "factory": "",
        "components_2": "",
        "recruitItems": "",
        "detail_JP": "",
        "detail": "",
    },
    "srg_umbryon": {
        "_id": 900501,
        "name_JP": "",
        "name": "Umbryon, Herald of Eternal Rot",
        "aka_JP": "",
        "aka": "",
        "idActor": "",
        "sort": "",
        "size": "",
        "_idRenderData": "",
        "tiles": 1502,
        "tiles_snow": "",
        "colorMod": "",
        "components": "",
        "defMat": "",
        "LV": 35,
        "chance": 0,
        "quality": 3,
        "hostility": "Enemy",
        "biome": "",
        "tag": "boss,noRandomProduct",
        "trait": "UniqueMonster",
        "race": "lich",
        "job": "warrior",
        "tactics": "",
        "aiIdle": "",
        "aiParam": "",
        "actCombat": "bolt_Darkness/50,miasma_Poison/30,ActDrainBlood/60",
        "mainElement": "Darkness",
        "elements": "life/100,resDarkness/20,END/10,WIL/10",
        "equip": "",
        "loot": "",
        "category": "",
        "filter": "",
        "gachaFilter": "",
        "tone": "",
        "actIdle": "",
        "lightData": "",
        "idExtra": "",
        "bio": "",
        "faith": "",
        "works": "",
        "hobbies": "",
        "idText": "",
        "moveAnime": "",
        "factory": "",
        "components_2": "",
        "recruitItems": "",
        "detail_JP": "",
        "detail": "",
    },
    "srg_solaris": {
        "_id": 900502,
        "name_JP": "",
        "name": "Solaris, Inferno of the Fallen Star",
        "aka_JP": "",
        "aka": "",
        "idActor": "",
        "sort": "",
        "size": "",
        "_idRenderData": "",
        "tiles": 1713,
        "tiles_snow": "",
        "colorMod": "",
        "components": "",
        "defMat": "",
        "LV": 40,
        "chance": 0,
        "quality": 3,
        "hostility": "Enemy",
        "biome": "",
        "tag": "boss,noRandomProduct",
        "trait": "UniqueMonster",
        "race": "spirit",
        "job": "warmage",
        "tactics": "",
        "aiIdle": "",
        "aiParam": "",
        "actCombat": "breathe_Fire/70,SpMeteor/10,bolt_Fire/50",
        "mainElement": "Fire",
        "elements": "resFire/30,life/80,MAG/15,mana/30",
        "equip": "",
        "loot": "",
        "category": "",
        "filter": "",
        "gachaFilter": "",
        "tone": "",
        "actIdle": "",
        "lightData": "",
        "idExtra": "",
        "bio": "",
        "faith": "",
        "works": "",
        "hobbies": "",
        "idText": "",
        "moveAnime": "",
        "factory": "",
        "components_2": "",
        "recruitItems": "",
        "detail_JP": "",
        "detail": "",
    },
    "srg_erevor": {
        "_id": 900503,
        "name_JP": "",
        "name": "Erevor, The Abyssal Maw",
        "aka_JP": "",
        "aka": "",
        "idActor": "",
        "sort": "",
        "size": "",
        "_idRenderData": "chara_L",
        "tiles": 104,
        "tiles_snow": "",
        "colorMod": "",
        "components": "",
        "defMat": "",
        "LV": 45,
        "chance": 0,
        "quality": 3,
        "hostility": "Enemy",
        "biome": "",
        "tag": "boss,noRandomProduct",
        "trait": "UniqueMonster",
        "race": "dragon",
        "job": "predator",
        "tactics": "",
        "aiIdle": "",
        "aiParam": "",
        "actCombat": "SpGravity/50,ActRush,breathe_Void/30",
        "mainElement": "Impact",
        "elements": "life/100,PDR/30,STR/15,END/15",
        "equip": "",
        "loot": "",
        "category": "",
        "filter": "",
        "gachaFilter": "",
        "tone": "",
        "actIdle": "",
        "lightData": "",
        "idExtra": "",
        "bio": "",
        "faith": "",
        "works": "",
        "hobbies": "",
        "idText": "",
        "moveAnime": "",
        "factory": "",
        "components_2": "",
        "recruitItems": "",
        "detail_JP": "",
        "detail": "",
    },
    "srg_quasarix": {
        "_id": 900504,
        "name_JP": "",
        "name": "Quasarix, Devourer of Light",
        "aka_JP": "",
        "aka": "",
        "idActor": "",
        "sort": "",
        "size": "",
        "_idRenderData": "",
        "tiles": 2317,
        "tiles_snow": "",
        "colorMod": "",
        "components": "",
        "defMat": "",
        "LV": 50,
        "chance": 0,
        "quality": 3,
        "hostility": "Enemy",
        "biome": "",
        "tag": "boss,noRandomProduct",
        "trait": "UniqueMonster",
        "race": "god",
        "job": "gunner",
        "tactics": "",
        "aiIdle": "",
        "aiParam": "",
        "actCombat": "SpSilence/50,ActGazeDim/30,bolt_Magic/60,SpBane/20",
        "mainElement": "Darkness",
        "elements": "resMagic/20,resDarkness/20,life/80,mana/50,MAG/15",
        "equip": "",
        "loot": "",
        "category": "",
        "filter": "",
        "gachaFilter": "",
        "tone": "",
        "actIdle": "",
        "lightData": "",
        "idExtra": "",
        "bio": "",
        "faith": "",
        "works": "",
        "hobbies": "",
        "idText": "",
        "moveAnime": "",
        "factory": "",
        "components_2": "",
        "recruitItems": "",
        "detail_JP": "",
        "detail": "",
    },
    "srg_archivist": {
        "_id": 900505,
        "name_JP": "",
        "name": "Astral Archivist",
        "aka_JP": "",
        "aka": "",
        "idActor": "",
        "sort": "",
        "size": "",
        "_idRenderData": "",
        "tiles": 1216,
        "tiles_snow": "",
        "colorMod": "",
        "components": "",
        "defMat": "",
        "LV": 30,
        "chance": 0,
        "quality": 4,
        "hostility": "Neutral",
        "biome": "",
        "tag": "neutral,noRandomProduct",
        "trait": "Archivist",
        "race": "elea",
        "job": "pianist",
        "tactics": "",
        "aiIdle": "",
        "aiParam": "",
        "actCombat": "SpHealHeavy/60,SpHero/50/pt",
        "mainElement": "",
        "elements": "featHealer/1,reading/10,MAG/10,WIL/10",
        "equip": "",
        "loot": "",
        "category": "",
        "filter": "",
        "gachaFilter": "",
        "tone": "",
        "actIdle": "",
        "lightData": "",
        "idExtra": "",
        "bio": "",
        "faith": "",
        "works": "",
        "hobbies": "",
        "idText": "",
        "moveAnime": "",
        "factory": "",
        "components_2": "",
        "recruitItems": "",
        "detail_JP": "",
        "detail": "",
    },
}


def find_rows(ws, col_map):
    rows = {}
    for row in range(4, ws.max_row + 1):
        item_id = ws.cell(row=row, column=col_map["id"]).value
        if isinstance(item_id, str):
            rows[item_id] = row
    return rows


def set_field(ws, row, col_map, key, value):
    col = col_map[key]
    cell = ws.cell(row=row, column=col)
    if value == "":
        value = None
    if cell.value != value:
        print(f"  FIX: {ws.cell(row=row, column=1).value} {key}: {cell.value!r} -> {value!r}")
        cell.value = value
        return 1
    return 0


def normalize_chara_typos(ws):
    rows = find_rows(ws, CHARA_COL)
    typo = rows.get("srg_ervor")
    correct = rows.get("srg_erevor")
    if typo is not None and correct is None:
        ws.cell(row=typo, column=CHARA_COL["id"]).value = "srg_erevor"
        print(f"  FIX: srg_ervor id: 'srg_ervor' -> 'srg_erevor' at row {typo}")
        return 1
    if typo is not None and correct is not None:
        ws.delete_rows(typo, 1)
        print(f"  DELETE: duplicate typo row srg_ervor at row {typo}")
        return 1
    return 0


def delete_forbidden_chara_rows(ws):
    rows = find_rows(ws, CHARA_COL)
    deleted = 0
    for chara_id in sorted(FORBIDDEN_CHARA_IDS):
        row = rows.get(chara_id)
        if row is not None:
            ws.delete_rows(row, 1)
            deleted += 1
            print(f"  DELETE: forbidden test Chara row {chara_id} at row {row}")
            rows = find_rows(ws, CHARA_COL)
    return deleted


def ensure_thing_rows(ws):
    rows = find_rows(ws, THING_COL)
    changed = 0
    added = 0

    for item_id, fields in EXPECTED_THINGS.items():
        row = rows.get(item_id)
        if row is None:
            row = ws.max_row + 1
            ws.cell(row=row, column=THING_COL["id"], value=item_id)
            rows[item_id] = row
            added += 1
            print(f"  ADD: {item_id} at row {row}")

        for key, value in fields.items():
            changed += set_field(ws, row, THING_COL, key, value)

        sort_value = ws.cell(row=row, column=THING_COL["sort_value"]).value
        if sort_value is not None and not isinstance(sort_value, int):
            print(f"  CLEAN: {item_id} numeric sort column {sort_value!r} -> None")
            ws.cell(row=row, column=THING_COL["sort_value"]).value = None
            changed += 1

    return added, changed


def ensure_chara_rows(ws):
    changed = delete_forbidden_chara_rows(ws)
    changed += normalize_chara_typos(ws)
    rows = find_rows(ws, CHARA_COL)
    added = 0

    for chara_id, fields in EXPECTED_CHARAS.items():
        row = rows.get(chara_id)
        if row is None:
            row = ws.max_row + 1
            ws.cell(row=row, column=CHARA_COL["id"], value=chara_id)
            rows[chara_id] = row
            added += 1
            print(f"  ADD: {chara_id} at row {row}")

        for key, value in fields.items():
            changed += set_field(ws, row, CHARA_COL, key, value)

    return added, changed


def main():
    wb = openpyxl.load_workbook(SOURCE_CARD)
    ws_thing = wb["Thing"]
    ws_chara = wb["Chara"]

    thing_added, thing_changed = ensure_thing_rows(ws_thing)
    chara_added, chara_changed = ensure_chara_rows(ws_chara)
    added = thing_added + chara_added
    changed = thing_changed + chara_changed

    if added or changed:
        wb.save(SOURCE_CARD)
        print(f"Saved SourceCard.xlsx ({added} added, {changed} fixed).")
    else:
        print("SourceCard.xlsx already matches expected Skyreader rows.")

    wb.close()
    normalize_shared_strings(SOURCE_CARD)
    print("Normalized shared strings in", SOURCE_CARD)


if __name__ == "__main__":
    main()
