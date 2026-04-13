import openpyxl
wb = openpyxl.load_workbook(r'D:\Steam\steamapps\workshop\content\2135150\3704602432\LangMod\EN\Adventurer_SourceCard.xlsx')
ws = wb.active
headers = [cell.value for cell in ws[1]]
rows = [[str(cell.value) for cell in row] for row in ws.iter_rows(min_row=2, max_row=4)]
for r in rows:
    print([f'{h}: {v}' for h, v in zip(headers, r) if h in ['id', 'tiles', '_idRenderData', 'idActor', 'idExtra']])
